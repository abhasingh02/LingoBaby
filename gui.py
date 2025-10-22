# gui.py
import tkinter as tk
from tkinter import messagebox
from excel_handler import add_new_sentence
from doc_handler import add_sentence_to_doc

def extract_new_words(sentence, existing_words):
    import re
    cleaned = re.sub(r'[\.\?!]', '', sentence)
    cleaned = re.sub(r'\[.*?\]', '', cleaned)  # remove text in [brackets]
    words = cleaned.split()
    new_words = [w for w in words if w.lower() not in existing_words]
    return new_words

def get_existing_words():
    import openpyxl
    try:
        wb = openpyxl.load_workbook("SmartVocabularyNotes.xlsx")
        sheet = wb["New Words"]
        return {str(sheet.cell(i, 2).value).lower() for i in range(2, sheet.max_row + 1)}
    except Exception:
        return set()

def on_submit():
    sentence = entry.get().strip()
    if sentence.lower() == "q":
        root.destroy()
        return

    existing_words = get_existing_words()
    new_words = extract_new_words(sentence, existing_words)

    if not new_words:
        messagebox.showinfo("Info", "No new words found.")
    else:
        add_new_sentence(new_words, sentence)
        add_sentence_to_doc(sentence, new_words)
        messagebox.showinfo("Success", f"Added: {', '.join(new_words)}")

    entry.delete(0, tk.END)

root = tk.Tk()
root.title("Smart Vocabulary")
root.geometry("400x200")
root.iconbitmap("icon.ico")

label = tk.Label(root, text="Enter Sentence:")
label.pack(pady=10)

entry = tk.Entry(root, width=50)
entry.pack(pady=5)

button = tk.Button(root, text="Submit", command=on_submit)
button.pack(pady=10)

exit_label = tk.Label(root, text="(Type Q to exit)")
exit_label.pack()

root.mainloop()
