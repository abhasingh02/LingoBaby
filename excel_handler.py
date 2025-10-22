# excel_handler.py
import os
from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from datetime import datetime
from helpers import clean_word_for_compare
from config import EXCEL_FILE, SHEET_NAME, TEMP_EXCEL

def init_workbook():
    """
    Load or create workbook and ensure the New Words sheet exists with header:
    ["No.", "New Words", "Sentence", "Date/Time"]
    Returns (wb, ws)
    """
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        if SHEET_NAME in wb.sheetnames:
            ws = wb[SHEET_NAME]
        else:
            ws = wb.create_sheet(SHEET_NAME)
            ws.append(["No.", "New Words", "Sentence", "Date/Time"])
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(["No.", "New Words", "Sentence", "Date/Time"])
    return wb, ws

def load_existing_words_set(ws):
    """
    Read column 2 (New Words) and return a set of cleaned lowercase words (for comparison).
    Only non-empty cells are included.
    """
    s = set()
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
        val = row[0]
        if val:
            # val might be a comma-separated string or a single word; treat as single word
            cleaned = clean_word_for_compare(str(val))
            if cleaned:
                s.add(cleaned)
    return s

def append_rows_for_sentence(ws, sentence, new_words):
    """
    Append one or more rows for the sentence.
    Each new word gets its own row (column 2 contains single new word).
    If new_words is empty, add one row with empty column 2.
    Column 1 (No.) will use ws.max_row as a running integer (header row counts).
    """
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # ws.max_row gives current last row index; we'll use that as starting No.
    next_no = ws.max_row
    if new_words:
        for w in new_words:
            ws.append([next_no, w, sentence, timestamp])
            next_no += 1
    else:
        ws.append([next_no, "", sentence, timestamp])

def save_workbook_safe(wb):
    """
    Save workbook to a temp file and atomically replace EXCEL_FILE to avoid lock issues.
    """
    # Save to temp then replace to minimize risk of file corruption
    wb.save(TEMP_EXCEL)
    wb.close()
    if os.path.exists(EXCEL_FILE):
        os.remove(EXCEL_FILE)
    os.replace(TEMP_EXCEL, EXCEL_FILE)
